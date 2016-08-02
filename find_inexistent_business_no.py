#coding=utf-8

# 查找不存在的商家编码
import sys
import openpyxl
import os
import user_input
import merge_sheet
import present

if __name__ == "__main__":
    present.use_des()
    print("开始处理\n")

    wb = merge_sheet.load_spec_file(user_input.file_name_original)

    # 输出数据的sheet名
    output_sheet_name = "不存在的商家编码"

    all_sheet_name = wb.get_sheet_names()
    output_sheet = None
    if output_sheet_name not in all_sheet_name:
        print("%s不存在, 创建它" % (output_sheet_name))
        output_sheet = wb.create_sheet(output_sheet_name)
    else:
        print("%s已经存在, 先删除它, 再创建它" % (output_sheet_name))
        wb.remove_sheet(wb[output_sheet_name])
        output_sheet = wb.create_sheet(output_sheet_name)

        """
        print("%s已经存在" % (output_sheet_name))
        output_sheet = wb[output_sheet_name]
        """

    # 计算出所有的商家编码
    all_business_no = {}
    cells = merge_sheet.cal_column_cell(wb["Sheet2"], "j", 2)
    for cell in cells:
        if cell.value not in all_business_no:
            all_business_no[cell.value] = 1
    all_business_no = all_business_no.keys()
    # print("\n\nall_business_no = %r\n\n" % (all_business_no))

    # 计算出当前存在的商家编码
    existent_business_no = {}
    cells = merge_sheet.cal_column_cell(wb["Sheet4"], "a", 2)
    for cell in cells:
        if cell.value not in existent_business_no:
            existent_business_no[cell.value] = 1
    existent_business_no = existent_business_no.keys()
    # print("\n\nexistent_business_no = %r\n\n" % (existent_business_no))

    # 计算出不存在的商家编码
    inexistent_business_no = {}
    for value in all_business_no:
        if value not in existent_business_no:
            if value not in inexistent_business_no:
                inexistent_business_no[value] = 1
                print("value = %r, 不存在" % (value))
    inexistent_business_no = inexistent_business_no.keys()
    print("\n\ninexistent_business_no = %r\n\n" % (inexistent_business_no))

    # 输出数据
    output_sheet.cell("a1").value = "不存在的商家编码"
    i = 2
    for value in inexistent_business_no:
        cell_to_do = output_sheet.cell("a%d" % (i))
        cell_to_do.value = value
        i += 1

    wb.save(user_input.file_name_original)
    present.after_process()

