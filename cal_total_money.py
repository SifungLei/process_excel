#coding=utf-8

# 计算商品金额
import sys
import openpyxl
import os
import user_input
import time
import present
import merge_sheet

if __name__ == "__main__":
    present.use_des()
    print("开始处理\n")

    wb = merge_sheet.load_spec_file(user_input.file_name_original)
    output_sheet = wb["Sheet1"]

    # 计算出符合条件的商家编码的行号
    row_to_do = {}
    cells = merge_sheet.cal_column_cell(output_sheet, "c", 2)
    for cell in cells:
        if isinstance(cell.value, str):
            # print("cell = %r, value = %r" % (cell, cell.value))
            if cell.value.count("发10双") > 0:
                print("cell = %r, value = %r, 满足条件" % (cell, cell.value))
                row_to_do[cell.row] = 1
                output_sheet.cell("f%d" % (cell.row)).value = 0.0
    row_to_do = row_to_do.keys()
    
    if len(row_to_do) == 0:
        print("找不到符合条件的数据")
        sys.exit(0)

    row_to_do = sorted(row_to_do)

    # K为快递单号，V为对应的购买数量
    express_no_buy_cnt = {}

    for row in row_to_do:
        express_no = output_sheet.cell("h%d" % (row)).value
        if express_no not in express_no_buy_cnt:
            express_no_buy_cnt[express_no] = 0
        express_no_buy_cnt[express_no] += output_sheet.cell("e%d" % (row)).value
    print("\n\nexpress_no_buy_cnt = %r\n\n" % (express_no_buy_cnt))

    row_done = {}
    for row in row_to_do:
        express_no = output_sheet.cell("h%d" % (row)).value

        if express_no in row_done:
            print("row = %d, express_no = %s, 已经处理过" % (row, express_no))
            output_sheet.cell("f%d" % (row)).value = 0.0
            continue

        row_done[express_no] = 1
        buy_cnt = express_no_buy_cnt[express_no]
        total_money = 0.0
        price = output_sheet.cell("d%d" % (row)).value
        if (buy_cnt % 2) == 0:
            # 偶数
            total_money = buy_cnt / 2 * price
        else:
            # 奇数
            total_money = (buy_cnt + 1) / 2 * price
        output_sheet.cell("f%d" % (row)).value = total_money
        print("row = %d, express_no = %s(对应的总的购买数量为%d), total_money = %f" % (row, express_no, buy_cnt, total_money))

    wb.save(user_input.file_name_original)
    present.after_process()

