#coding=utf-8

# 合并N个sheet的数据为一个sheet
import sys
import openpyxl
import os
import user_input
import time
import present

# 计算某一列的单元格
def cal_column_cell(sheet, col_str, from_row):
    """
    from_row为从第几行开始处理
    """
    index = from_row - 1
    ret = sheet.columns[openpyxl.cell.column_index_from_string(col_str) - 1][index:]
    return ret

# 单元格，通过VLOOKUP来计算
def cal_cell_by_vlookup(from_sheet, from_col_k, from_col_v, from_row, to_sheet, to_col_k, to_col_v):
    # 查找源数据
    all_data = {} # 所有的源数据
    column_valid_cell = cal_column_cell(from_sheet, from_col_k, from_row)
    for cell in column_valid_cell:
        value = cell.value
        if value not in all_data:
            all_data[value] = from_sheet.cell("%s%d" % (from_col_v, cell.row)).value
    # print("all_data= %r" % (all_data))

    # 引用源数据
    column_valid_cell = cal_column_cell(to_sheet, to_col_k, from_row)
    for cell in column_valid_cell:
        value = cell.value
        # output_cell(cell)
        cell_to_do = to_sheet.cell("%s%d" % (to_col_v, cell.row))
        if value not in all_data:
            print("calling cal_cell_by_vlookup, cell = %r, value = %r, 不存在于(%r)" % (cell, value, from_sheet))
            wb.save(user_input.file_name_original)
            sys.exit(0)
            continue
        cell_to_do.value = all_data[value]

# 打印单元格的值
def output_cell(cell):
    print("sheet_name = %s, coordinate = %s, value = %r" % (cell.parent.title, cell.coordinate, cell.value))

# 加载指定的文件
def load_spec_file(file_name):
    present.check_file_exist(file_name)

    print("准备加载%s" % (file_name))
    wb = openpyxl.load_workbook(file_name, data_only=True)
    if wb == None:
        print("加载%s失败" % (file_name))
        sys.exit(0)
    print("加载%s完毕" % (file_name))
    return wb

# 复制一个表的列数据到另一个表
def copy_sheet_column(from_sheet, from_col, from_row, to_sheet, to_col, des):
    print("calling copy_sheet_column, 正在处理%s, 复制一个表的列数据到另一个表, from_sheet = %r, from_col = %s, from_row = %d, to_sheet = %r, to_col = %s" % (des, from_sheet, from_col, from_row, to_sheet, to_col))
    cells = cal_column_cell(from_sheet, from_col, from_row)
    i = from_row
    j = 1
    for cell in cells:
        cell_to_do = to_sheet.cell("%s%d" % (to_col, i))
        cell_to_do.value = cell.value
        """
        if j == 1:
            output_cell(cell)
        """
        i += 1
        j += 1

# 输出数据的sheet
def cal_output_sheet(wb):
    output_sheet_name = "Sheet1"
    all_sheet_name = wb.get_sheet_names()
    output_sheet = None
    if output_sheet_name not in all_sheet_name:
        print("%s不存在, 创建它" % (output_sheet_name))
        output_sheet = wb.create_sheet(output_sheet_name, 0)
    else:
        """
        print("%s已经存在, 先删除它, 再创建它" % (output_sheet_name))
        wb.remove_sheet(wb[output_sheet_name])
        output_sheet = wb.create_sheet(output_sheet_name, 0)
        """

        print("%s已经存在" % (output_sheet_name))
        output_sheet = wb[output_sheet_name]

    print("calling cal_output_sheet, output_sheet = %r" % (output_sheet))
    return output_sheet

if __name__ == "__main__":
    present.use_des()
    print("请确保源数据的第1行都是标题名，python会从第2行开始处理数据\n")
    print("开始合并数据\n")

    wb = load_spec_file(user_input.file_name_original)
    output_sheet = cal_output_sheet(wb)

    # 第1行，列名
    column_name = ["订单编号","店铺名称","商家编码","价格","购买数量","商品金额合计","快递公司","快递单号","补快递差价","赠送费","加礼盒费","邮费","卖家备注","买家留言","收货地址","付款时间","订单成本"]
    column_cnt = len(column_name)
    for i in range(0, column_cnt):
        cell_to_do = output_sheet.cell("%s1" % (openpyxl.cell.get_column_letter(i + 1)))
        cell_to_do.value = column_name[i]

    # 订单编号
    copy_sheet_column(wb["Sheet2"], "a", 2, output_sheet, "a", "订单编号")

    row_cnt = len(wb["Sheet2"].columns[0])
    print("row_cnt = %r" % row_cnt)
    # sys.exit(0)

    # 店铺名称
    shop_name = wb["Sheet3"].columns[openpyxl.cell.column_index_from_string("aa") - 1][1].value
    if shop_name == None or len(shop_name) == 0:
        print("error,店铺名称为空")
        sys.exit(0)
    for i in range(2, row_cnt + 1):
        cell_to_do = output_sheet.cell("b%d" % (i))
        cell_to_do.value = shop_name

    # 商家编码
    copy_sheet_column(wb["Sheet2"], "j", 2, output_sheet, "c", "商家编码")

    # 价格
    cal_cell_by_vlookup(wb["Sheet4"], "a", "b", 2, output_sheet, "c", "d")

    # 购买数量
    copy_sheet_column(wb["Sheet2"], "d", 2, output_sheet, "e", "购买数量")

    wb.save(user_input.file_name_original)

    # 商品金额
    for i in range(2, row_cnt + 1):
        x = "f%d" % (i)
        cell_to_do = output_sheet.cell(x)
        # print("正在处理商品金额, 单元格坐标 = %s" % x)
        price = float(output_sheet.cell("d%d" % (i)).value)
        buy_cnt = float(output_sheet.cell("e%d" % (i)).value)
        cell_to_do.value =  price * buy_cnt

    # 快递公司
    cal_cell_by_vlookup(wb["Sheet3"], "a", "w", 2, output_sheet, "a", "g")

    # 快递单号
    cal_cell_by_vlookup(wb["Sheet3"], "a", "v", 2, output_sheet, "a", "h")

    # 邮费
    cells = cal_column_cell(output_sheet, "h", 2)
    postage_order_no = []
    for cell in cells:
        value = cell.value # 快递单号
        price = user_input.postage # 邮费
        cell_to_do = output_sheet.cell("l%d" % (cell.row))
        if value not in postage_order_no:
            print("行号 = %d, 快递单号 = %r, 第一次出现, 邮费 = %r" % (cell.row, value, price))
            cell_to_do.value = price 
            postage_order_no.append(value)
        else:
            print("行号 = %d, 快递单号 = %r, 不是第一次出现了, 邮费 = %r" % (cell.row, value, 0))
            cell_to_do.value = 0

    # 卖家备注
    cal_cell_by_vlookup(wb["Sheet3"], "a", "x", 2, output_sheet, "a", "m")

    # 买家留言
    cal_cell_by_vlookup(wb["Sheet3"], "a", "l", 2, output_sheet, "a", "n")

    # 收货地址
    cal_cell_by_vlookup(wb["Sheet3"], "a", "n", 2, output_sheet, "a", "o")

    # 付款时间
    cal_cell_by_vlookup(wb["Sheet3"], "a", "s", 2, output_sheet, "a", "p")

    # 订单成本
    cells = cal_column_cell(output_sheet, "f", 2)
    for cell in cells:
        cell_to_do = output_sheet.cell("q%d" % (cell.row))
        cell_to_do.value = 0.0

        value = cell.value
        if value != None:
            cell_to_do.value += float(value)

        value = output_sheet.cell("i%d" % (cell.row)).value
        if value != None:
            cell_to_do.value += float(value)

        value = output_sheet.cell("j%d" % (cell.row)).value
        if value != None:
            cell_to_do.value += float(value)

        value = output_sheet.cell("k%d" % (cell.row)).value
        if value != None:
            cell_to_do.value += float(value)

        value = output_sheet.cell("l%d" % (cell.row)).value
        if value != None:
            cell_to_do.value += float(value)

    wb.save(user_input.file_name_original)
    present.after_process()


